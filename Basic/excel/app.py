import openpyxl as xl

wb = xl.load_workbook("Book.xlsx")
sheet = wb['Sheet1']
print(sheet.max_row)
cell = sheet["A1"]
print(cell.value)
cell = sheet.cell(1, 2)
print(cell.value)

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value + .05
    corrected_price_cell = sheet.cell(row, 5)
    corrected_price_cell.value = corrected_price

wb.save('Book.xlsx')
